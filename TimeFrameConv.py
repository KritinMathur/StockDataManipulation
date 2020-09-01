import openpyxl as pyxl

wb = pyxl.load_workbook('NIFTY25JUN2010000PE.xlsx')
ws = wb.worksheets[0]

wb_new = pyxl.Workbook()
ws_new = wb_new.worksheets[0]

ws_new.title = ws.title


ws_new['A1'] = ws['A1'].value
ws_new['B1'] = ws['B1'].value
ws_new['C1'] = ws['C1'].value
ws_new['D1'] = ws['D1'].value
ws_new['E1'] = ws['E1'].value
ws_new['F1'] = ws['F1'].value
ws_new['G1'] = ws['G1'].value
ws_new['H1'] = ws['H1'].value

TICKER = ws['A2'].value
DATE = ws['B2'].value

rowcount = ws.max_row
count = 1

for i in range(0,int((rowcount+1)/15)):

        highli = []
        lowli = []
        volli = []
        for j in range(1,16):
                highli.append(ws['E'+str(i*15 + j +1)].value)
                lowli.append(ws['F'+str(i*15 + j +1)].value)
                volli.append(ws['H'+str(i*15 + j +1)].value)
                
                if(j == 1):
                        openval = ws['D'+str(i*15 + j +1)].value
                        timestamp = ws['C'+str(i*15 + j +1)].value

                if(j == 15):
                        closeval = ws['G'+str(i*15 + j +1)].value

                if(j == 15):
                        ws_new['D'+str(i+2)] = openval
                        ws_new['G'+str(i+2)] = closeval
                        ws_new['C'+str(i+2)] = timestamp
                        ws_new['E'+str(i+2)] = max(highli)
                        ws_new['F'+str(i+2)] = min(lowli)
                        ws_new['H'+str(i+2)] = sum(volli)
                        ws_new['A'+str(i+2)] = TICKER
                        ws_new['B'+str(i+2)] = DATE
                
wb_new.save('NIFTY25JUN2010000PE.xlsx')
