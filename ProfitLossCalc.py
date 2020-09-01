import openpyxl as pyxl

wb = pyxl.load_workbook('NIFTY25JUN2010000PE.xlsx')
ws = wb.worksheets[0]
rowcount = ws.max_row

short = False
prof = 0
for i in range(3,rowcount+1):
        prelo = ws['F'+str(i-1)].value
        curlo = ws['F'+str(i)].value

        prehi = ws['E'+str(i-1)].value
        curhi = ws['E'+str(i)].value

        curtime = ws['C'+str(i)].value
        
        if( curlo < prelo and short == False):
                short = True
                entryval = ws['F'+str(i)].value
                continue

        if((prehi < curhi or (curtime.hour==15 and curtime.minute >= 15))and short == True):
                exitval = ws['E'+str(i)].value
                prof += entryval - exitval
                short = False

                continue


if (prof < 0):
        print("LOSS :- ",prof)
else:
        print("PROFIT :- ",prof)

input()
